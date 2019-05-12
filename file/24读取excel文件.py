"""读取 excel 文件，推荐使用 openpyxl。

操作 excel 文件的话，建议直接使用 pandas 。
"""

import xlrd
from openpyxl import load_workbook


def use_xlrd(path):
    """不建议使用"""

    workbook = xlrd.open_workbook(path)
    # sheet索引从0开始
    sheet1 = workbook.sheet_by_index(0)
    # 获取行数
    print(sheet1.nrows)
    # 获取列数
    print(sheet1.ncols)
    # 获取第4行内容
    rows = sheet1.row_values(3)
    # 获取第1列内容
    cols = sheet1.col_values(0)
    # 获取第 2 行，第 1 列的内容
    cell = sheet1.cell(2, 1)
    # 转换为字典
    d = dict(zip(sheet1.col_values(0), sheet1.col_values(1)))

    # 迭代行
    for row_idx in range(1, sheet1.nrows):
        row = sheet1.row_values(row_idx)


def use_openpyxl(path):
    """使用 openpyxl 读取 excel"""

    # 读取，`read_only` 可以提高性能
    wb = load_workbook(path, read_only=True)

    # 返回一个 sheet 对象的 list
    sheet_list = wb.worksheets
    print(sheet_list)

    # 返回一个 sheet 名字的 list
    sheet_name_list = wb.sheetnames
    print(sheet_name_list)

    # 读取 sheet 页，根据名字
    # sheet = wb['测试页']

    # 获取当前正在使用的 sheet 页
    sheet = wb.active

    # 获取 sheet 属性
    print(sheet.title)       # 当前 sheet 名字
    print(sheet.max_column)  # 当前 sheet 最大列数
    print(sheet.max_row)       # 当前 sheet 最大行数

    # 获取某一行，注意列数从 1 开始
    # 第一行是列名
    print([c.value for c in sheet[1]])
    print([c.value for c in sheet[2]])

    # 可以用迭代器迭代 1 整行
    print(next(sheet.rows))

    # 迭代指定行，限制列成员，注意这里 min_row 从 1 开始
    # 并且第一行是列名
    # 不指定 max_row，则迭代所有
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=2, max_col=3):
        print([c.value for c in row])

    # 迭代指定列，限制行成员
    # for col in sheet.iter_cols(min_row=1, min_col=1, max_row=3, max_col=2):
    #     print(col)

    # 选定单个单元格
    row1 = sheet[2]
    cell1 = row1[2]
    print(cell1)

    # 单元格的属性
    print(cell1.value)  # 值


def main() -> None:
    path = '/home/zzzzer/Documents/data/数据集/最新电池数据/data1_temp2.xlsx'
    # use_xlrd(path)
    use_openpyxl(path)


if __name__ == '__main__':
    main()
